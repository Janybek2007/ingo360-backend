from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook


def _get_nested_value(data: dict[str, Any], path: str) -> Any:
    current: Any = data
    for part in path.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


"""
def _get_nested_value(data: Any, path: str) -> Any:
    current = data
    for part in path.split("."):
        if current is None:
            return None
        if isinstance(current, dict):
            current = current.get(part)
        else:
            current = getattr(current, part, None)
    return current
"""


def _format_template(template: str, data: dict[str, Any]) -> str:
    formatted = template
    for key in data.keys():
        formatted = formatted.replace(f"{{{key}}}", str(data.get(key) or ""))
    return formatted.strip()


def _apply_custom_map(rule: dict[str, str], row: dict[str, Any]) -> str:
    # rule: {"is_admin": "administrator", "is_operator": "operator"}
    for flag_field, label in rule.items():
        value = row.get(flag_field)
        if value is True or value == 1:
            return label
    return ""


def build_export_row_values(
    *,
    row: dict[str, Any],
    headers: list[str],
    header_map: dict[str, str],
    fields_map: dict[str, str] | None = None,
    boolean_map: dict[str, list[str]] | None = None,
    custom_map: dict[str, dict[str, str]] | None = None,
) -> list[Any]:
    fields_map = fields_map or {}
    boolean_map = boolean_map or {}
    custom_map = custom_map or {}

    values: list[Any] = []
    for key in headers:
        if key in custom_map:
            value = _apply_custom_map(custom_map[key], row)
        elif key in fields_map:
            value = _format_template(fields_map[key], row)
        elif "." in key:
            value = _get_nested_value(row, key)
        else:
            value = row.get(key)

        if key in boolean_map and value is not None:
            normalized = 1 if value is True or value == 1 else 0
            labels = boolean_map[key]
            if 0 <= normalized < len(labels):
                value = labels[normalized]

        values.append("" if value is None else value)

    return values


def export_excel_to_file(
    *,
    rows: list[dict[str, Any]],
    header_map: dict[str, str],
    output_path: str,
    fields_map: dict[str, str] | None = None,
    boolean_map: dict[str, list[str]] | None = None,
    custom_map: dict[str, dict[str, str]] | None = None,
) -> str:
    headers = list(header_map.keys())
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append([header_map[key] for key in headers])

    for row in rows:
        ws.append(
            build_export_row_values(
                row=row,
                headers=headers,
                header_map=header_map,
                fields_map=fields_map,
                boolean_map=boolean_map,
                custom_map=custom_map,
            )
        )

    wb.save(str(path))
    wb.close()
    return str(path)


def export_excel(
    *,
    rows: list[dict[str, Any]],
    header_map: dict[str, str],
    fields_map: dict[str, str] | None = None,
    boolean_map: dict[str, list[str]] | None = None,
    custom_map: dict[str, dict[str, str]] | None = None,
) -> bytes:
    headers = list(header_map.keys())

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append([header_map[key] for key in headers])

    for row in rows:
        ws.append(
            build_export_row_values(
                row=row,
                headers=headers,
                header_map=header_map,
                fields_map=fields_map,
                boolean_map=boolean_map,
                custom_map=custom_map,
            )
        )

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer.read()
