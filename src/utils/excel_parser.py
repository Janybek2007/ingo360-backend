from io import BytesIO
from tempfile import SpooledTemporaryFile
from typing import Iterator

import pandas as pd
from fastapi import UploadFile
from openpyxl import load_workbook


async def save_upload_to_temp(
    file: UploadFile, chunk_size: int = 1024 * 1024
) -> SpooledTemporaryFile:
    temp = SpooledTemporaryFile(max_size=50 * 1024 * 1024)
    while True:
        chunk = await file.read(chunk_size)
        if not chunk:
            break
        temp.write(chunk)
    temp.seek(0)
    return temp


def iter_excel_records(
    file_obj: SpooledTemporaryFile, read_as_str: bool = False
) -> Iterator[tuple[int, dict]]:
    file_obj.seek(0)
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        workbook.close()
        return iter(())

    normalized_headers = [
        str(h).strip().lower() if h is not None else "" for h in headers
    ]
    row_index = 1

    for row in rows:
        row_index += 1
        if not row or all(cell is None for cell in row):
            continue

        record: dict = {}
        for header, cell in zip(normalized_headers, row):
            if not header:
                continue
            value = None if cell == "-" else cell
            if isinstance(value, str):
                value = value.strip()
            if read_as_str and value is not None:
                value = str(value)
            record[header] = value

        yield row_index, record

    workbook.close()


async def parse_excel_file(file: UploadFile, read_as_str: bool = False) -> list[dict]:
    content = await file.read()

    if read_as_str:
        df = pd.read_excel(BytesIO(content), sheet_name=0, dtype=str)
    else:
        df = pd.read_excel(BytesIO(content), sheet_name=0)

    df = df.dropna(how="all")
    df = df.where(pd.notna(df), None)
    df = df.replace("-", None)

    df.columns = df.columns.str.strip().str.lower()

    str_cols = df.select_dtypes(include="object").columns
    df[str_cols] = df[str_cols].apply(lambda col: col.str.strip())

    return df.to_dict("records")
